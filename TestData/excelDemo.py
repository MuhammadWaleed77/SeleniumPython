import openpyxl
book =openpyxl.load_workbook("/Users/numuqa/Downloads/MyLearnings/PythonDemo.xlsx") #loades excel sheet
sheet =book.active # select active sheet Sheet1
Dict = {}
cell =sheet.cell(row=1, column=2) #get row and cell
print(cell.value)
sheet.cell(row=2, column=2).value = "Waleed" #write into sheet

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A5'].value) #A column 5th row

for i in range(1,sheet.max_row+1):  # to get rows give +1 so that it doesn't exclude the last row
    if sheet.cell(row =i,column=1).value == "Testcase2":  # get onlyTestCase2 values

        for j in range(2,sheet.max_column+1):#to get columns Use 1 if need to show TestCase2 (col value)
            #Dict["lastname"]="shetty
            print(sheet.cell(row=i, column=j).value)
            # use dictionary as in tc we r calling data using dictionary
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value


print(Dict)







